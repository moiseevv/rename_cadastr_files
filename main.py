import os
import shutil
from openpyxl import Workbook
from openpyxl import load_workbook as lwb

path_dir = r'C:\Users\vmoiseev\Downloads\files'
path_xl = r'D:\1\доки\исполнение\ГЛОБАЛЬНАЯ ПО ВСЕМ ВЫПИСКАМ\Авдеев_2022_09_22\Сопоставление Запрос-Номер.xlsx'

#Номер строки с которой начинается xl
first_str = 2

def write_file(name,str,n_f):
    name = name+'.txt'
    with open(name,'a') as file:
        print(f"На строке {str}   в файле   {n_f}",file = file)



if __name__ == '__main__':

    lb = lwb(path_xl)
    ws = lb.active

    max_xl = ws.max_row
    print("Всего строк", max_xl)

    for i in range(first_str,max_xl+1):
        kadastr = ws.cell(i,2).value
        name_file = ws.cell(i,1).value

        z_1 = path_dir+'\\'+name_file+'.pdf'

        name_clear = str(kadastr).replace(':','_')
        print(name_clear)
        z_2 = path_dir+'\\'+name_clear+'.pdf'

        if os.path.exists(z_1):
            os.rename(z_1,z_2)
            print(f"{i} Переименован {z_1}   в   {z_2}")
            write_file('success', i, name_file)
        else:
            print("ОШИБКА - ", z_1)
            write_file('errors',i,name_file)


